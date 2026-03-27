from fastapi import FastAPI, File, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from typing import Optional
import openpyxl
import io
import re
import json
import os
from datetime import datetime, date

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

GATE_ORDER = ["PCI", "FeG", "CSG", "CG", "DG", "FDG", "FIG", "RG", "EG"]
DATA_FILE = "qtp_data.json"


# ---- JSON DB helpers ----

def read_db():
    empty = {"tasks": [], "engineers": [], "project_gates": {}, "gate_deviations": {}, "audit_log": []}
    if not os.path.exists(DATA_FILE):
        return empty
    try:
        with open(DATA_FILE, "r", encoding="utf-8") as f:
            content = f.read().strip()
        if not content:
            return empty
        db = json.loads(content)
    except (json.JSONDecodeError, ValueError):
        # File is corrupt - back it up and start fresh
        import shutil
        try:
            shutil.copy(DATA_FILE, DATA_FILE + ".bak")
        except Exception:
            pass
        db = empty
    db.setdefault("gate_deviations", {})
    db.setdefault("audit_log", [])
    return db

def write_db(db):
    import tempfile
    dir_name = os.path.dirname(os.path.abspath(DATA_FILE)) or '.'
    with tempfile.NamedTemporaryFile(mode='w', suffix='.tmp', dir=dir_name, delete=False, encoding='utf-8') as f:
        json.dump(db, f, indent=2)
        tmp_path = f.name
    os.replace(tmp_path, DATA_FILE)


# ---- Models ----

class Task(BaseModel):
    project_id: str
    item: str
    quality_task: str
    gate: str
    quality_engineer: str
    responsible_engineer: str
    expected_status: str
    actual_status: str
    rating: str
    followup_week: str
    planned_start: Optional[str] = None
    planned_end: Optional[str] = None
    completion_level: str
    comments: Optional[str] = ""
    next_gate: Optional[str] = ""
    deviation: Optional[str] = "No"
    deviation_text: Optional[str] = ""


class DeleteTask(BaseModel):
    project_id: str
    item: str
    quality_task: str
    gate: str


class EngineerName(BaseModel):
    name: str


class AdvanceGate(BaseModel):
    project_id: str
    new_gate: str


class GateDeviation(BaseModel):
    project_id: str
    gate: str
    deviated: bool
    reason: Optional[str] = ""
    changed_by: Optional[str] = ""


class AuditEntry(BaseModel):
    project_id: str
    gate: Optional[str] = ""
    item: Optional[str] = ""
    quality_task: Optional[str] = ""
    field_changed: str
    old_value: Optional[str] = ""
    new_value: Optional[str] = ""
    changed_by: str
    changed_at: str


# ---------- SAVE TASK ----------
@app.post("/save-task")
def save_task(task: Task):
    db = read_db()
    tasks = db["tasks"]

    # Check if ANY rows exist for this project + item + quality_task
    existing = [t for t in tasks if t["project_id"] == task.project_id
                and t["item"] == task.item and t["quality_task"] == task.quality_task]

    if not existing:
        # FIRST TIME: Insert 9 rows, one per gate
        for g in GATE_ORDER:
            if g == task.gate:
                tasks.append({
                    "project_id": task.project_id, "item": task.item,
                    "quality_task": task.quality_task, "gate": g,
                    "quality_engineer": task.quality_engineer,
                    "responsible_engineer": task.responsible_engineer,
                    "expected_status": task.expected_status,
                    "actual_status": task.actual_status,
                    "rating": task.rating, "followup_week": task.followup_week,
                    "planned_start": task.planned_start or "",
                    "planned_end": task.planned_end or "",
                    "completion_level": task.completion_level,
                    "comments": task.comments or "",
                    "next_gate": task.next_gate or "",
                    "deviation": task.deviation or "No",
                    "deviation_text": task.deviation_text or "",
                    "current_gate": db["project_gates"].get(task.project_id, "PCI")
                })
            else:
                tasks.append({
                    "project_id": task.project_id, "item": task.item,
                    "quality_task": task.quality_task, "gate": g,
                    "quality_engineer": task.quality_engineer,
                    "responsible_engineer": task.responsible_engineer,
                    "expected_status": "", "actual_status": "", "rating": "",
                    "followup_week": "", "planned_start": "", "planned_end": "",
                    "completion_level": "", "comments": "",
                    "next_gate": task.next_gate or "",
                    "deviation": "No", "deviation_text": "",
                    "current_gate": db["project_gates"].get(task.project_id, "PCI")
                })
    else:
        # UPDATE: find the specific gate row and update it
        for t in tasks:
            if (t["project_id"] == task.project_id and t["item"] == task.item
                    and t["quality_task"] == task.quality_task and t["gate"] == task.gate):
                t.update({
                    "quality_engineer": task.quality_engineer,
                    "responsible_engineer": task.responsible_engineer,
                    "expected_status": task.expected_status,
                    "actual_status": task.actual_status,
                    "rating": task.rating,
                    "followup_week": task.followup_week,
                    "planned_start": task.planned_start or "",
                    "planned_end": task.planned_end or "",
                    "completion_level": task.completion_level,
                    "comments": task.comments or "",
                    "next_gate": task.next_gate or "",
                    "deviation": task.deviation or "No",
                    "deviation_text": task.deviation_text or ""
                })
        # Update shared fields across all gates for this task
        for t in tasks:
            if (t["project_id"] == task.project_id and t["item"] == task.item
                    and t["quality_task"] == task.quality_task):
                t["quality_engineer"] = task.quality_engineer
                t["responsible_engineer"] = task.responsible_engineer
                t["next_gate"] = task.next_gate or ""

    write_db(db)
    return {"status": "saved"}


# ---------- LOAD PROJECT ----------
@app.get("/project/{project_id}")
def get_project(project_id: str, gate: Optional[str] = None):
    db = read_db()
    tasks = db["tasks"]

    if gate:
        data = [t for t in tasks if t["project_id"] == project_id and t["gate"] == gate]
    else:
        data = [t for t in tasks if t["project_id"] == project_id]

    current_gate = db["project_gates"].get(project_id, "PCI")
    return {"data": data, "current_gate": current_gate}


# ---------- DELETE TASK ----------
@app.post("/delete-task")
def delete_task(req: DeleteTask):
    db = read_db()
    db["tasks"] = [t for t in db["tasks"] if not (
        t["project_id"] == req.project_id and t["item"] == req.item
        and t["quality_task"] == req.quality_task and t["gate"] == req.gate
    )]
    write_db(db)
    return {"status": "deleted"}


# ---------- ENGINEERS ----------
@app.get("/engineers")
def get_engineers():
    db = read_db()
    return {"engineers": sorted(db.get("engineers", []))}


@app.post("/save-engineer")
def save_engineer(eng: EngineerName):
    db = read_db()
    if eng.name not in db["engineers"]:
        db["engineers"].append(eng.name)
        write_db(db)
    return {"status": "saved"}


# ---------- ADVANCE GATE ----------
@app.post("/advance-gate")
def advance_gate(req: AdvanceGate):
    db = read_db()
    db["project_gates"][req.project_id] = req.new_gate
    for t in db["tasks"]:
        if t["project_id"] == req.project_id:
            t["current_gate"] = req.new_gate
    write_db(db)
    return {"status": "advanced", "current_gate": req.new_gate}


# ---------- GATE DEVIATIONS ----------
@app.post("/save-gate-deviation")
def save_gate_deviation(req: GateDeviation):
    db = read_db()
    if req.project_id not in db["gate_deviations"]:
        db["gate_deviations"][req.project_id] = {}
    if req.deviated:
        db["gate_deviations"][req.project_id][req.gate] = {
            "deviated_at": datetime.now().isoformat(),
            "reason": req.reason or "",
            "reverted_at": None,
            "changed_by": req.changed_by or ""
        }
    else:
        db["gate_deviations"][req.project_id].pop(req.gate, None)
    write_db(db)
    return {"status": "saved"}


@app.get("/project-deviations/{project_id}")
def get_project_deviations(project_id: str):
    db = read_db()
    deviations = db["gate_deviations"].get(project_id, {})
    active = {g: info for g, info in deviations.items() if not info.get("reverted_at")}
    return {"deviations": active}


# ---------- AUDIT LOG ----------
@app.post("/save-audit")
def save_audit_entry(entry: AuditEntry):
    db = read_db()
    db["audit_log"].append({
        "project_id": entry.project_id,
        "gate": entry.gate or "",
        "item": entry.item or "",
        "quality_task": entry.quality_task or "",
        "field_changed": entry.field_changed,
        "old_value": entry.old_value or "",
        "new_value": entry.new_value or "",
        "changed_by": entry.changed_by,
        "changed_at": entry.changed_at
    })
    write_db(db)
    return {"status": "saved"}


@app.get("/audit-log/{project_id}")
def get_audit_log(project_id: str):
    db = read_db()
    logs = [e for e in db["audit_log"] if e["project_id"] == project_id]
    logs.sort(key=lambda x: x["changed_at"], reverse=True)
    return {"logs": logs[:100]}


# ---------- OVERDUE TASKS ----------
@app.get("/overdue-tasks")
def get_overdue_tasks():
    today = date.today().isoformat()
    db = read_db()
    overdue = []
    seen = set()
    for t in db["tasks"]:
        end = t.get("planned_end", "")
        if not end or not end.strip():
            continue
        status = t.get("actual_status", "").lower()
        if end < today and status not in ("completed", "n/a"):
            key = (t["project_id"], t["item"], t["quality_task"], t["gate"])
            if key not in seen:
                seen.add(key)
                overdue.append({
                    "project_id": t["project_id"],
                    "gate": t["gate"],
                    "item": t["item"],
                    "task": t["quality_task"],
                    "planned_end": end,
                    "actual_status": t.get("actual_status", "")
                })
    return {"overdue": overdue}


# ---------- ALL PROJECTS ----------
@app.get("/all-projects")
def get_all_projects():
    db = read_db()
    project_ids = list(db["project_gates"].keys())

    results = []
    for pid in project_ids:
        tasks = [t for t in db["tasks"] if t["project_id"] == pid]
        current_gate = db["project_gates"].get(pid, "PCI")

        # Only count tasks for the current gate
        gate_tasks = [t for t in tasks if t["gate"] == current_gate]
        total = len(gate_tasks)
        green  = sum(1 for t in gate_tasks if t.get("rating","").lower() == "green")
        yellow = sum(1 for t in gate_tasks if t.get("rating","").lower() == "yellow")
        red    = sum(1 for t in gate_tasks if t.get("rating","").lower() == "red")

        # Completion: tasks with actual_status == Completed across all gates up to current
        gate_idx = GATE_ORDER.index(current_gate) if current_gate in GATE_ORDER else 0
        done_tasks = [t for t in tasks if t["gate"] in GATE_ORDER[:gate_idx+1]
                      and t.get("actual_status","").lower() == "completed"]
        all_filled = [t for t in tasks if t["gate"] in GATE_ORDER[:gate_idx+1]
                      and t.get("actual_status","")]
        completion_pct = round(len(done_tasks) / len(all_filled) * 100) if all_filled else 0

        # Deviations = active gate deviations for this project
        dev_gates = db["gate_deviations"].get(pid, {})
        deviations = sum(1 for v in dev_gates.values() if not v.get("reverted_at"))

        results.append({
            "project_id": pid,
            "current_gate": current_gate,
            "total_tasks": total,
            "green": green, "yellow": yellow, "red": red,
            "completion_pct": completion_pct,
            "deviations": deviations,
        })

    results.sort(key=lambda x: x["project_id"])
    return {"projects": results}


# ---------- UPLOAD QTP EXCEL ----------
COL = {
    "item": 4, "quality_task": 5, "responsible_engineer": 7,
    "quality_engineer": 8, "expected_status": 9, "completion_level": 10,
    "actual_status": 11, "rating": 12, "deviation": 13,
    "followup_week": 14, "planned_start": 15, "planned_end": 16, "comments": 18,
}
HEADER_ROW = 7
DATA_START = 8


def _clean(val):
    if val is None:
        return ""
    s = str(val).strip()
    if s.lower() in ("none", "nan"):
        return ""
    return s


def _parse_date(val):
    if val is None:
        return None
    from datetime import datetime, date
    if isinstance(val, (datetime, date)):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    if not s or s.lower() in ("none", "nan", "tbd", "n/a", "na", "-"):
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%Y/%m/%d", "%d.%m.%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def _find_qtp_sheet(wb):
    for name in wb.sheetnames:
        if "qtp" in name.lower():
            return wb[name]
    return wb.worksheets[0]


@app.post("/upload-qtp")
async def upload_qtp(file: UploadFile = File(...)):
    try:
        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
        ws = _find_qtp_sheet(wb)

        project_id = _clean(ws.cell(2, 3).value) or "UNKNOWN"
        project_id = project_id.replace("/", "-").replace("  ", " ").strip()

        current_gate = _clean(ws.cell(2, 13).value) or "PCI"
        current_gate = current_gate.strip()
        if current_gate not in GATE_ORDER:
            for g in GATE_ORDER:
                if g.lower() == current_gate.lower():
                    current_gate = g
                    break

        parsed_tasks = []
        last_item = ""
        last_resp = ""
        last_qe = ""

        for r in range(DATA_START, ws.max_row + 1):
            qt = _clean(ws.cell(r, COL["quality_task"]).value)
            if not qt:
                continue
            item_val = _clean(ws.cell(r, COL["item"]).value)
            if item_val:
                last_item = item_val
            resp_val = _clean(ws.cell(r, COL["responsible_engineer"]).value)
            if resp_val:
                last_resp = resp_val
            qe_val = _clean(ws.cell(r, COL["quality_engineer"]).value)
            if qe_val:
                last_qe = qe_val

            dev = _clean(ws.cell(r, COL["deviation"]).value)
            parsed_tasks.append({
                "item": last_item or "Default",
                "quality_task": qt,
                "responsible_engineer": last_resp.replace("\n", ", "),
                "quality_engineer": last_qe.replace("\n", ", "),
                "expected_status": _clean(ws.cell(r, COL["expected_status"]).value),
                "actual_status": _clean(ws.cell(r, COL["actual_status"]).value),
                "rating": _clean(ws.cell(r, COL["rating"]).value),
                "completion_level": _clean(ws.cell(r, COL["completion_level"]).value),
                "deviation": "No",
                "deviation_text": dev if dev.lower() not in ("no", "na", "") else "",
                "followup_week": _clean(ws.cell(r, COL["followup_week"]).value),
                "planned_start": _parse_date(ws.cell(r, COL["planned_start"]).value),
                "planned_end": _parse_date(ws.cell(r, COL["planned_end"]).value),
                "comments": _clean(ws.cell(r, COL["comments"]).value),
            })

        if not parsed_tasks:
            return {"status": "error", "detail": "No task rows found in the QTP sheet."}

        db = read_db()
        db["project_gates"][project_id] = current_gate

        # Collect engineers
        all_engineers = set(db.get("engineers", []))
        for t in parsed_tasks:
            for name in re.split(r"[,\n]", t["responsible_engineer"]):
                name = name.strip()
                if name:
                    all_engineers.add(name)
            for name in re.split(r"[,\n]", t["quality_engineer"]):
                name = name.strip()
                if name:
                    all_engineers.add(name)
        db["engineers"] = list(all_engineers)

        inserted = 0
        for t in parsed_tasks:
            existing = [x for x in db["tasks"] if x["project_id"] == project_id
                        and x["item"] == t["item"] and x["quality_task"] == t["quality_task"]]
            if existing:
                for x in db["tasks"]:
                    if (x["project_id"] == project_id and x["item"] == t["item"]
                            and x["quality_task"] == t["quality_task"] and x["gate"] == current_gate):
                        x.update({
                            "quality_engineer": t["quality_engineer"],
                            "responsible_engineer": t["responsible_engineer"],
                            "expected_status": t["expected_status"],
                            "actual_status": t["actual_status"],
                            "rating": t["rating"],
                            "followup_week": t["followup_week"],
                            "planned_start": t["planned_start"] or "",
                            "planned_end": t["planned_end"] or "",
                            "completion_level": t["completion_level"],
                            "comments": t["comments"],
                            "deviation": t["deviation"],
                            "deviation_text": t["deviation_text"],
                            "current_gate": current_gate
                        })
            else:
                for g in GATE_ORDER:
                    if g == current_gate:
                        db["tasks"].append({
                            "project_id": project_id, "item": t["item"],
                            "quality_task": t["quality_task"], "gate": g,
                            "quality_engineer": t["quality_engineer"],
                            "responsible_engineer": t["responsible_engineer"],
                            "expected_status": t["expected_status"],
                            "actual_status": t["actual_status"],
                            "rating": t["rating"],
                            "followup_week": t["followup_week"],
                            "planned_start": t["planned_start"] or "",
                            "planned_end": t["planned_end"] or "",
                            "completion_level": t["completion_level"],
                            "comments": t["comments"],
                            "next_gate": "", "deviation": t["deviation"],
                            "deviation_text": t["deviation_text"],
                            "current_gate": current_gate
                        })
                    else:
                        db["tasks"].append({
                            "project_id": project_id, "item": t["item"],
                            "quality_task": t["quality_task"], "gate": g,
                            "quality_engineer": t["quality_engineer"],
                            "responsible_engineer": t["responsible_engineer"],
                            "expected_status": "", "actual_status": "", "rating": "",
                            "followup_week": "", "planned_start": "", "planned_end": "",
                            "completion_level": "", "comments": "", "next_gate": "",
                            "deviation": "No", "deviation_text": "",
                            "current_gate": current_gate
                        })
                inserted += 1

        write_db(db)
        return {
            "status": "ok",
            "project_id": project_id,
            "current_gate": current_gate,
            "tasks_imported": len(parsed_tasks),
            "tasks_new": inserted,
            "tasks_updated": len(parsed_tasks) - inserted,
            "items": list(set(t["item"] for t in parsed_tasks))
        }

    except Exception as e:
        return {"status": "error", "detail": str(e)}


# ---------- SERVE FRONTEND ----------
app.mount("/", StaticFiles(directory="static", html=True), name="static")
